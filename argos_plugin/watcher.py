"""Spec-07 (#71): the watcher — document catalog, extraction & freshness.

The doc tier's engine room: scan → catalog → hot policy → extraction →
freshness lifecycle. Closes #68 (D4 trust/freshness fields) and #70
(audit seed) on landing.

Three tiers:
  1. **Catalog pass** — every file, near-zero cost, no LLM. Stat-only
     walk (path/size/mtime), classify new/changed/moved/unchanged/deleted.
  2. **Extraction pass** — hot docs only, facts with evidence into Argos
     via the normal write path (proposal/review gates, supersession, egress).
  3. **Raw files stay in place** — Argos stores provenance and paths,
     never copies.

Identity: `file_id` = SHA-256 of raw bytes (dedup, alias detection,
content-edit detection); `extract_hash` = SHA-256 of the extraction
input (PDF text layer / sheet values + headers). Re-extraction fires
only when `extract_hash` changes — Excel resaves are no-ops.

Backward compatible: **no watcher config = zero behaviour change.**

No new deps (hashlib stdlib; PyPDF2 + openpyxl + python-docx already in
the venv). Deterministic, no LLM calls in the catalog pass.
"""
from __future__ import annotations

import hashlib
import json
import logging
import os
import re
import stat
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# D1 — Stable doc identity: content hash, not path
# ---------------------------------------------------------------------------

_CHUNK_SIZE = 65536  # 64KB chunks for streamed hashing


def hash_file(path: str | Path) -> str | None:
    """Streamed SHA-256 of file content. Returns None on error (locked,
    permission, missing). Never raises — the scan pass must not fail."""
    try:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            while True:
                chunk = f.read(_CHUNK_SIZE)
                if not chunk:
                    break
                h.update(chunk)
        return h.hexdigest()
    except (PermissionError, OSError, IOError) as exc:
        logger.debug("hash_file skip %s: %s", path, exc)
        return None


def hash_extraction_input(text: str) -> str:
    """SHA-256 of the extraction input (PDF text layer / sheet values +
    headers). This is the re-extract gate — Excel resaves that don't
    change the data produce the same hash → no re-extraction."""
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# D2 — file_catalog table is created in store_core.py
# ---------------------------------------------------------------------------

# Doc type classification from file extension.
_DOC_TYPES: Dict[str, str] = {
    ".pdf": "pdf",
    ".xlsx": "xlsx",
    ".xls": "xlsx",
    ".csv": "csv",
    ".docx": "docx",
    ".doc": "docx",
}

# Exclude patterns (temp files, lock files, OS metadata).
_EXCLUDE_PATTERNS = [
    re.compile(r"^~\$", re.IGNORECASE),  # Office lock files
    re.compile(r"^\.tmp$", re.IGNORECASE),
    re.compile(r"^thumbs\.db$", re.IGNORECASE),
    re.compile(r"^\.DS_Store$", re.IGNORECASE),
    re.compile(r"\.tmp$", re.IGNORECASE),
]


def classify_doc_type(path: str | Path) -> str | None:
    """Return the doc type (pdf/xlsx/csv/docx) or None if unsupported."""
    ext = Path(path).suffix.lower()
    return _DOC_TYPES.get(ext)


def is_excluded(name: str) -> bool:
    """Check if a filename matches an exclude pattern."""
    for pat in _EXCLUDE_PATTERNS:
        if pat.search(name):
            return True
    return False


# ---------------------------------------------------------------------------
# D3 — The scan loop
# ---------------------------------------------------------------------------

# Classification of a file relative to the catalog.
#   new      — not in catalog
#   changed  — in catalog, content hash differs (version bump)
#   moved    — in catalog (same hash), different path (alias update)
#   unchanged — in catalog, same hash, same path
#   deleted  — in catalog, path no longer exists (tombstone)


def scan_pass(
    roots: List[str | Path],
    catalog: Dict[str, Dict[str, Any]],
) -> Dict[str, List[Dict[str, Any]]]:
    """Walk scan roots and classify files against the catalog.

    Args:
        roots: directories to scan (local + UNC shares).
        catalog: current catalog state — {file_id: {paths: [...], ...}}.
            Also indexed by path for move detection.

    Returns:
        Dict with keys 'new', 'changed', 'moved', 'unchanged', 'deleted',
        'skipped' — each a list of file info dicts.

    The pass never fails. Locked/open files are skipped + retried next pass.
    """
    # Build a path → file_id index from the catalog for move detection.
    path_to_id: Dict[str, str] = {}
    for file_id, entry in catalog.items():
        for p in entry.get("paths", []):
            path_to_id[p] = file_id

    result: Dict[str, List[Dict[str, Any]]] = {
        "new": [],
        "changed": [],
        "moved": [],
        "unchanged": [],
        "deleted": [],
        "skipped": [],
    }

    seen_paths: set[str] = set()

    for root in roots:
        root_path = Path(root)
        if not root_path.exists():
            logger.warning("scan root %s does not exist", root)
            continue
        try:
            for dirpath, dirnames, filenames in os.walk(root_path):
                # Filter excluded filenames.
                filenames = [
                    f for f in filenames if not is_excluded(f)
                ]
                for fname in filenames:
                    fpath = str(Path(dirpath) / fname)
                    seen_paths.add(fpath)
                    doc_type = classify_doc_type(fpath)
                    if doc_type is None:
                        continue  # unsupported format, skip
                    try:
                        st = os.stat(fpath)
                    except (PermissionError, OSError) as exc:
                        result["skipped"].append({
                            "path": fpath, "reason": str(exc),
                        })
                        continue
                    # Hash the file (streamed). Locked files return None.
                    file_id = hash_file(fpath)
                    if file_id is None:
                        result["skipped"].append({
                            "path": fpath, "reason": "locked or unreadable",
                        })
                        continue
                    # Spec-09 (#112): form-level identity. Compute the layout
                    # fingerprint for new/changed files only (file_id not in
                    # catalog). Unchanged/moved files already have it in the
                    # catalog (same content → same fingerprint). LLM-free,
                    # deterministic. Lazy import to keep the module importable
                    # without the fingerprint module in edge test setups.
                    layout_family = None
                    if file_id not in catalog:
                        try:
                            from layout_fingerprint import compute_layout_family
                            layout_family = compute_layout_family(fpath, doc_type)
                        except Exception as exc:
                            logger.debug("layout fingerprint skip %s: %s", fpath, exc)
                            layout_family = None
                    info = {
                        "path": fpath,
                        "file_id": file_id,
                        "size": st.st_size,
                        "mtime": datetime.fromtimestamp(
                            st.st_mtime, tz=timezone.utc
                        ).isoformat(),
                        "doc_type": doc_type,
                        "layout_family": layout_family,
                    }
                    if file_id in catalog:
                        # Known content — check if it's at the same path.
                        cat_entry = catalog[file_id]
                        if fpath in cat_entry.get("paths", []):
                            result["unchanged"].append(info)
                        else:
                            # Same content, different path = moved (alias).
                            result["moved"].append(info)
                    elif fpath in path_to_id:
                        # Path was known but content changed.
                        result["changed"].append(info)
                    else:
                        result["new"].append(info)
        except (PermissionError, OSError) as exc:
            logger.warning("scan walk failed for %s: %s", root, exc)

    # Detect deleted: catalog paths not seen in this scan.
    for file_id, entry in catalog.items():
        for p in entry.get("paths", []):
            if p not in seen_paths and entry.get("status") == "active":
                result["deleted"].append({
                    "path": p,
                    "file_id": file_id,
                })

    return result


# ---------------------------------------------------------------------------
# Spec-09 (#112): layout-family classification of scan results
# ---------------------------------------------------------------------------


def classify_scan_by_layout(
    scan_result: Dict[str, List[Dict[str, Any]]],
    known_families: Dict[str, int],
) -> Dict[str, List[Dict[str, Any]]]:
    """Classify the 'new' and 'changed' buckets by layout family.

    Returns a dict with keys 'known', 'novel', 'unknown' — each a list of
    file info dicts (the same dicts from scan_result['new'/'changed'],
    with an added 'layout_class' field).

    'known'   — fingerprint matches an existing family in the catalog.
    'novel'   — fingerprint not seen in the catalog → surface for full
                extraction + human labelling (never silently skipped).
    'unknown' — fingerprint could not be computed (unreadable/unsupported);
                the file is still processed normally, just no form-level
                shortcut.

    The known-family short-circuit (skip re-fingerprinting / family lookup
    on subsequent passes) is subordinate to content: the caller must still
    apply the content-level re-extract gate (``extract_hash``) for 'changed'
    files. See ``layout_fingerprint.should_short_circuit_extraction``.
    """
    try:
        from layout_fingerprint import classify_layout
    except ImportError:
        classify_layout = None  # type: ignore[assignment]
    out: Dict[str, List[Dict[str, Any]]] = {
        "known": [],
        "novel": [],
        "unknown": [],
    }
    for bucket in ("new", "changed"):
        for info in scan_result.get(bucket, []):
            fam = info.get("layout_family")
            if classify_layout is not None:
                cls = classify_layout(fam, known_families)
            else:
                cls = "unknown" if fam is None else ("known" if fam in known_families else "novel")
            info_with_cls = dict(info)
            info_with_cls["layout_class"] = cls
            out[cls].append(info_with_cls)
    return out


def surface_novel_layouts(
    scan_result: Dict[str, List[Dict[str, Any]]],
    known_families: Dict[str, int],
) -> List[Dict[str, Any]]:
    """Return the list of novel-layout files that must surface for full
    extraction + human labelling.

    Per the issue acceptance: 'novel layouts always surface for full
    extraction + human labelling (never silently skipped).'
    """
    return classify_scan_by_layout(scan_result, known_families)["novel"]


# ---------------------------------------------------------------------------
# D3 — Heuristic one-line descriptions
# ---------------------------------------------------------------------------


def heuristic_description(
    path: str | Path,
    doc_type: str,
    *,
    first_page_text: str | None = None,
    sheet_names: List[str] | None = None,
    first_lines: str | None = None,
) -> str:
    """Deterministic one-liner from filename + folder + content hints.

    No LLM calls. The description is built from:
    - The filename (always available)
    - The parent folder name (often the client or category)
    - First-page text (PDF) / sheet names (XLSX) / first lines (CSV)

    LLM descriptions are only used when heuristic confidence is low
    (separate pass, daily cap — not implemented here).
    """
    p = Path(path)
    parts: List[str] = []
    # Folder context.
    parent = p.parent.name
    if parent and parent not in (".", "/", "\\"):
        parts.append(parent)
    # Filename without extension.
    stem = p.stem
    parts.append(stem)
    # Content hint.
    if doc_type == "pdf" and first_page_text:
        # First non-empty line, truncated.
        for line in first_page_text.split("\n"):
            line = line.strip()
            if line:
                parts.append(line[:80])
                break
    elif doc_type == "xlsx" and sheet_names:
        parts.append(", ".join(sheet_names[:3]))
    elif doc_type == "csv" and first_lines:
        for line in first_lines.split("\n"):
            line = line.strip()
            if line:
                parts.append(line[:80])
                break
    elif doc_type == "docx" and first_page_text:
        for line in first_page_text.split("\n"):
            line = line.strip()
            if line:
                parts.append(line[:80])
                break
    return " — ".join(parts)[:200]


# ---------------------------------------------------------------------------
# D4 — Hot policy
# ---------------------------------------------------------------------------

# Default hot policy: recency window in days, usage threshold, type defaults.
_DEFAULT_RECENCY_DAYS = 365  # current tax/financial year
_DEFAULT_USAGE_THRESHOLD = 3  # N touches → auto-promote
_TYPE_DEFAULTS: Dict[str, bool] = {
    "pdf": False,
    "xlsx": True,  # spreadsheets are often master data
    "csv": False,
    "docx": False,
}


def is_hot(
    info: Dict[str, Any],
    *,
    now: str | None = None,
    recency_days: int = _DEFAULT_RECENCY_DAYS,
    usage_threshold: int = _DEFAULT_USAGE_THRESHOLD,
    type_defaults: Dict[str, bool] | None = None,
) -> Tuple[bool, str]:
    """Evaluate the hot policy. Returns (is_hot, reason).

    Precedence: pin > usage > type-default > recency.
    """
    # Pin override wins.
    if info.get("pinned"):
        return True, "pinned"
    # Usage: N touches → auto-promote.
    touch_count = int(info.get("touch_count", 0))
    if touch_count >= usage_threshold:
        return True, f"usage({touch_count})"
    # Type default.
    td = type_defaults or _TYPE_DEFAULTS
    doc_type = info.get("doc_type", "")
    if td.get(doc_type, False):
        return True, f"type-default({doc_type})"
    # Recency: mtime within the window.
    mtime = info.get("mtime")
    if mtime:
        try:
            mt = datetime.fromisoformat(mtime.replace("Z", "+00:00"))
            ref = datetime.now(timezone.utc)
            if now:
                ref = datetime.fromisoformat(now.replace("Z", "+00:00"))
            age_days = (ref - mt).days
            if age_days <= recency_days:
                return True, f"recency({age_days}d)"
        except (ValueError, TypeError):
            pass
    return False, "not-hot"


# ---------------------------------------------------------------------------
# D6 — Filename-pattern conflict surfacing
# ---------------------------------------------------------------------------

# Correction markers in filenames.
_CORRECTION_MARKERS = re.compile(
    r"(?i)(corrected|revised|\(\s*\d+\s*\)|v\d+|rev\b|amended|updated)"
)


def has_correction_marker(filename: str) -> bool:
    """Check if a filename carries a correction marker (CORRECTED, REVISED,
    (2), v2, etc.). Used for cross-doc conflict surfacing — zero LLM."""
    return bool(_CORRECTION_MARKERS.search(filename))


def extract_doc_type_label(filename: str) -> str:
    """Extract a document type label from a filename for conflict matching.

    Strips correction markers and extensions to get the base name that
    two versions of the same document would share.
    """
    stem = Path(filename).stem
    # Remove correction markers to get the base name.
    base = _CORRECTION_MARKERS.sub("", stem).strip(" -_")
    return base or stem


# ---------------------------------------------------------------------------
# D5 — Extraction input preparation (text extraction from files)
# ---------------------------------------------------------------------------


def extract_text_pdf(path: str | Path) -> Tuple[str, str]:
    """Extract text from a PDF. Returns (text, method).

    method is 'text' for born-digital PDFs, 'ocr' for scanned (fallback).
    Uses PyPDF2 for text extraction. If the text layer is empty or very
    sparse, returns method='ocr' to flag the document as needing OCR
    (the actual OCR is a separate pass, not implemented here).
    """
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(str(path))
        pages: List[str] = []
        for page in reader.pages:
            try:
                text = page.extract_text() or ""
            except Exception:
                text = ""
            pages.append(text)
        full_text = "\n".join(pages).strip()
        # Heuristic: if we got very little text, it's likely a scan.
        if len(full_text) < 50 and len(reader.pages) > 0:
            return full_text, "ocr"
        return full_text, "text"
    except Exception as exc:
        logger.debug("PDF text extraction failed for %s: %s", path, exc)
        return "", "text"


def extract_text_xlsx(path: str | Path) -> Tuple[str, List[str]]:
    """Extract values from an Excel file. Returns (text, sheet_names).

    The text is the extraction input: sheet values + headers encoded
    per row (the 26/8 rule). This is what gets hashed as extract_hash.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        lines: List[str] = []
        for sname in sheet_names:
            ws = wb[sname]
            lines.append(f"[Sheet: {sname}]")
            for row in ws.iter_rows(values_only=True):
                # Encode headers + values per row.
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    lines.append("\t".join(cells))
        wb.close()
        return "\n".join(lines), sheet_names
    except Exception as exc:
        logger.debug("XLSX extraction failed for %s: %s", path, exc)
        return "", []


def extract_text_csv(path: str | Path) -> str:
    """Extract text from a CSV file. Returns the raw text content."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()
    except Exception as exc:
        logger.debug("CSV extraction failed for %s: %s", path, exc)
        return ""


def extract_text_docx(path: str | Path) -> str:
    """Extract plain text from a DOCX file."""
    try:
        import docx
        doc = docx.Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as exc:
        logger.debug("DOCX extraction failed for %s: %s", path, exc)
        return ""


def prepare_extraction_input(
    path: str | Path,
    doc_type: str,
) -> Tuple[str, str, str]:
    """Prepare the extraction input for a document.

    Returns (text, method, extract_hash).
    - text: the extraction input (PDF text layer / sheet values / CSV content)
    - method: 'text' / 'ocr' / 'excel'
    - extract_hash: SHA-256 of the text (the re-extract gate)
    """
    if doc_type == "pdf":
        text, method = extract_text_pdf(path)
        return text, method, hash_extraction_input(text)
    elif doc_type in ("xlsx", "xls"):
        text, sheet_names = extract_text_xlsx(path)
        # Include sheet names in the description hint.
        return text, "excel", hash_extraction_input(text)
    elif doc_type == "csv":
        text = extract_text_csv(path)
        return text, "text", hash_extraction_input(text)
    elif doc_type == "docx":
        text = extract_text_docx(path)
        return text, "text", hash_extraction_input(text)
    return "", "text", hash_extraction_input("")


# ---------------------------------------------------------------------------
# D4 — verified_state lifecycle
# ---------------------------------------------------------------------------

# Verified states for document-sourced facts.
VERIFIED_CURRENT = "current"
VERIFIED_UNVERIFIED = "unverified"
VERIFIED_STALE = "stale"
VERIFIED_INVALIDATED = "invalidated"


def verified_state_on_version_bump() -> str:
    """Old facts become stale when the source document gets a new version."""
    return VERIFIED_STALE


def verified_state_on_tombstone() -> str:
    """Facts become invalidated when the source document is deleted."""
    return VERIFIED_INVALIDATED


def verified_state_on_ocr_numeric() -> str:
    """OCR-sourced numeric facts are born unverified (flagged, lower
    confidence). A principal verify action flips to current."""
    return VERIFIED_UNVERIFIED


def verified_state_on_verify() -> str:
    """A principal verify action flips unverified → current."""
    return VERIFIED_CURRENT


# ---------------------------------------------------------------------------
# D5 — Doc-fact extraction (LLM pass, uses extraction_llm_model/provider)
# ---------------------------------------------------------------------------

_DOC_FACT_SYSTEM_PROMPT = """You are a document fact extractor for a legal/accounting practice memory system.

Extract factual statements from the document text below. Each fact must be:
- A standalone statement (readable without the surrounding context)
- Attributable to this document (not general knowledge)
- Numeric values must be exact (transcribe digits, don't round)

Output JSON: a list of objects with keys:
  "content": the fact statement (string)
  "category": one of "personal_fact", "event", "goal", "context_note"
  "source_loc": page/sheet/row reference if discernible (string, or null)
  "confidence": 0.0-1.0 (your confidence in the extraction accuracy)

Output ONLY the JSON list. No commentary."""


def extract_doc_facts_llm(
    text: str,
    *,
    model: str = "",
    provider: str = "",
    timeout: float = 30.0,
) -> List[Dict[str, Any]]:
    """Extract facts from document text using the LLM.

    Uses the extraction-specific model/provider (Spec-08 #72). Falls
    back to the auxiliary client default when empty. Never raises —
    the extraction pass must not crash the watcher.

    Returns a list of fact dicts, or empty list on any failure.
    """
    if not text or len(text.strip()) < 50:
        return []
    try:
        from egress import gate as _egress_gate
        if not _egress_gate("watcher_extraction", text):
            return []
    except Exception:
        pass  # egress gate unavailable — fail soft

    try:
        from agent.auxiliary_client import call_llm
    except ImportError:
        logger.debug("LLM unavailable for doc-fact extraction: auxiliary_client not importable")
        return []
    except Exception:
        return []

    messages = [
        {"role": "system", "content": _DOC_FACT_SYSTEM_PROMPT},
        {"role": "user", "content": text[:8000]},  # cap input length
    ]
    try:
        response = call_llm(
            task="doc_fact_extraction",
            messages=messages,
            temperature=0.0,
            max_tokens=2000,
            timeout=timeout,
            model=model or None,
            provider=provider or None,
        )
    except Exception as exc:
        logger.debug("Doc-fact LLM call failed: %s", exc)
        return []
    if response is None:
        return []
    try:
        raw = response.choices[0].message.content
    except (AttributeError, IndexError, KeyError):
        return []
    try:
        facts = json.loads(raw)
        if isinstance(facts, list):
            return facts
    except (json.JSONDecodeError, TypeError):
        pass
    return []


def extract_facts_from_doc(
    path: str | Path,
    doc_type: str,
    *,
    extraction_llm_model: str = "",
    extraction_llm_provider: str = "",
) -> Tuple[List[Dict[str, Any]], str, str, str]:
    """Full extraction pipeline for a single document.

    1. Prepare extraction input (text from PDF/XLSX/CSV/DOCX).
    2. Hash the extraction input (extract_hash — the re-extract gate).
    3. Extract facts via LLM (using extraction-specific model/provider).

    Returns (facts, extract_hash, method, text).
    - facts: list of fact dicts from the LLM
    - extract_hash: SHA-256 of the extraction input
    - method: 'text' / 'ocr' / 'excel'
    - text: the extraction input (for debugging/audit)
    """
    text, method, extract_hash = prepare_extraction_input(path, doc_type)
    if not text.strip():
        return [], extract_hash, method, text
    facts = extract_doc_facts_llm(
        text,
        model=extraction_llm_model,
        provider=extraction_llm_provider,
    )
    return facts, extract_hash, method, text
